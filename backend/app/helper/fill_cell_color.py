from openpyxl.styles import PatternFill

FILL_MATCH = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")     # green
FILL_CASE  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")     # yellow
FILL_MISMATCH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # red
FILL_MISSING  = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # grey


def apply_cmp_color(cell, value):
    if value == "MATCH":
        cell.fill = FILL_MATCH
    elif value == "CASE_DIFF":
        cell.fill = FILL_CASE
    elif value == "MISMATCH":
        cell.fill = FILL_MISMATCH
    elif value == "MISSING":
        cell.fill = FILL_MISSING